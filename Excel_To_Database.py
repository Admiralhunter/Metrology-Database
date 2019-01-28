import pandas as pd
import openpyxl
import win32com.client as win32
import time
import datetime
import sqlite3
import numpy as np


def main():

    # ensure this link is to the database for the metrology data else the program will not work properly
    database = "C:/Users/Hpalcich/PycharmProjects/MeterologyDatabase/MetrologyData.sqlite"
    databaseconn = "file:" + database + "?mode=rw"


    # ensure this link is to the excel sheet that gives the location's of the metrology data
    # use / instead of \. ( \ is used to signify tabs in a string and thus its just easier to use /)
    link = "O:/Quality/LinkBetweenQCandDatabsLocstest.xlsx"

    # Checks to see if master database for Metrology exists
    # if the .db doesnt exist then the file will be created.
    try:
        #conn = sqlite3.connect('file:MetrologyData.sqlite?mode=rw', uri=True)
        conn = sqlite3.connect(databaseconn, uri=True)
        c = conn.cursor()
        c.close()
    except sqlite3.OperationalError:

        conn = sqlite3.connect(database)
        c = conn.cursor()
        command = "CREATE TABLE users (Company text, Parent_Company text);"

        c.execute(command)
        conn.commit()
        c.close()
    finally:
        conn = sqlite3.connect(database)





    try:
        DBlinks = pd.read_excel(link, header=None)
    except FileNotFoundError:

        text = "This is an automated email alerting you that the Excel_To_Database.py file can" \
               " not find the" + link + " file. Most likely someone has moved or deleted it."
        recipient = "hpalcich@MicroTekFinishing.com"
        subject = "Automated Message: Excel_To_Database.py Workbook Open Error"
        send_mail_via_com(text, subject, recipient)

    wbk = openpyxl.load_workbook(link)
    wks = wbk['Sheet1']


    if DBlinks.shape[1] == 2:   # checks to see if any links had been move, if they have then fix this excel sheet
        for i in range(0, DBlinks.shape[0]):    # goes through each row to check if the link had been moved
            if not isinstance(DBlinks.iat[i, 1], float):     # checks to see if the value is an empty cell in excel
                DBlinks.iat[i, 0] = DBlinks.iat[i, 1]

            wks.cell(row=i + 1, column=1).value = DBlinks.iat[i, 0]

        DBlinks.drop(axis=1, columns=1, inplace=True)  # remove the second column to return sheet back to normal

    # write the column back into the excel file and close it
    failure = True
    attempts = 0
    while failure:  # keep trying to save and close workbook
        try:
            wbk.save(link)
            wbk.close()
            failure = False
        except PermissionError:  # if error arises from workbook being opened wait and try again
            if attempts < 2:
                time.sleep(180)  # if an errors occurs, wait 3 mins to try again.
                attempts = attempts + 1
                continue
            else:  # if errors are persistent then just stop and send an email letting someone know

                text = "This is an automated email alerting you that the Excel_To_Database.py file can" \
                       " not save the LinkBetweenQCandDatabsLocstest.xlsx file. Most likely someone has it open."
                recipient = "hpalcich@MicroTekFinishing.com"
                subject = "Automated Message: Excel_To_Database.py Workbook Save Error"
                send_mail_via_com(text, subject, recipient)
                failure = False

    for i in range(0, len(DBlinks)):

        print('Reading Data from ' + DBlinks.iat[i, 0])

        try:  # Tries to read each file and gather the data
            Data = pd.read_excel(DBlinks.iat[i, 0], header=None)
            Data.replace(0, np.nan, inplace= True)

        except FileNotFoundError:
            text = DBlinks.iat[i, 0] + " can not be found and opened. Check to ensure that the file exists" \
                                       " at the correct location."
            recipient = "hpalcich@MicroTekFinishing.com"
            subject = "Automated Message: Excel_To_Database.py Workbook Open Error"
            send_mail_via_com(text, subject, recipient)

        for j in range(0, len(Data.index)):  # Gets the first row that the data is inputted
            job = str(Data.iat[j, 0])
            firstrow = job.find("-")
            if firstrow != -1:
                firstrow = j
                break

        headers = [Data.iloc[0, :]]  # Give headers to our Dataframe for easier data manipulation
        for j in range(1, firstrow):
            for x in range(0, Data.shape[1]):
                if x > Data.shape[1] and j > 1:  # Necessary for the Average, Mins and Max's
                    break
                else:
                    if str(Data.iat[j, x]) != "nan":
                        headers[0][x] = str(headers[0][x]) + '_' + str(Data.iat[j, x])
                        headers[0][x].replace(" ", "_")


        Data.columns = headers

        Data.drop(Data.index[range(0, firstrow)], axis=0, inplace=True)  # remove columns that don't contain data


        if firstrow == -1:  # If no new data is found, then continue on to the next part
            continue

        # gets name of job and part to check if part table has been created or not
        table = DBlinks[0][i]
        tablename = table.split(".xlsx")
        table = tablename[0]
        tablename = table.split("\\")
        tablename = tablename[-1]

        # Checks to see if master database for Metrology exists
        # if the .db doesnt exist then the file will be created.
        try:
            conn = sqlite3.connect(databaseconn, uri=True)
            c = conn.cursor()
            c.close()
        except sqlite3.OperationalError:

            conn = sqlite3.connect(database)
            c = conn.cursor()

            c.close()
        finally:
            conn = sqlite3.connect(database)


        check_for_table(tablename, Data.columns.values.tolist(), database)  # Checks if table for part exists, else creates one
        insert_data(tablename, Data, DBlinks, i, database)  # Inputs data into table for Database
    testtime = datetime.datetime.now()

    #if testtime.hour == 9  and  testtime.minute < 10:  # this works, just pick a time for it to run
    if True:
        data_integrity_check(database)

######################################
# BELOW THIS LINE ARE ALL FUNCTIONS THAT ARE CALLED OUT IN THE PROGRAM


# This function is to run through all the tables in the database and ensure there are the correct amount of data
# points in each column and that values are positive (will need to change that for rsk)


def data_integrity_check(database):
    conn = sqlite3.connect(database)
    c = conn.cursor()
    command = "SELECT name FROM sqlite_master WHERE type='table';"
    c.execute(command)
    tables = c.fetchall()  # connect to database and get names of all the tables in it

    initial_text = "There is an integrity problem with the Metrology Database."
    text = ""
    final_text = ""

    for i in range(0, len(tables)):  # loop through all the tables and get the data and column names and make a df
        command = "SELECT * From " + tables[i][0] + ";"

        c.execute(command)
        collection = c.fetchall()
        data = pd.DataFrame(collection)

        command = "PRAGMA table_info(" + tables[i][0] + ")"
        c.execute(command)
        columns = c.fetchall()


        column_names = []
        for j in range(0, len(columns)):
            column_names.append(columns[j][1])

        column_names = pd.Series(column_names)
        data.columns = column_names

        col_null_totals = data.count()  # get count of all Null, None, Nan of each column
        col_null = col_null_totals[(col_null_totals != data.shape[0]) & (col_null_totals != 0)]  # check to see if any
        # column has a count of missing values that are not either 0 (not missing any data) or equal to the number
        # of rows (some columns never get data due to us unable to measure the spot)
        detected_null = False  # boolean to check whether null values exist
        if len(col_null) > 0:  # if there are Null values, create a string with all columns and how many values are missing
            detected_null = True

            col_null = data.shape[0] - col_null
            text = text + " For Table: " + tables[i][0] + ", there" \
            " are missing or extra data in the following columns followed by total rows affected for each columns. "
            columns_with_errors = col_null.index
            for x in range(0, len(col_null)):
                if col_null[x] > 0:

                    text = text + columns_with_errors[x] + ": " + str(col_null[x]) + " columns with missing data. "

                else:
                    col_null[x] = col_null[x]*-1
                    text = text + columns_with_errors[x] + ": " + str(col_null[x]) + " columns with extra rows of data. "

        table_text = " For Table: " + tables[i][0] + ", there" \
                                                            " are negative values in the following columns"
        column_text = ". "
        detected_negative = False
        for column in data.columns[4:]:  # Loops through columns to check if any values are negative
            column_values = data[column].astype(np.float32)
            if column.find('Rsk') != -1 or column.find('Rku') != -1:  # rsk and rku can be negative so we have to check for this
                continue
            negative_values = any(column_values < 0)
            if negative_values:
                detected_negative = True
                column_text = column_text + column + ", "

        column_text = column_text[:-2] + "."



        if detected_negative and detected_null:
            detected_negative = False
            detected_null = False
            final_text = initial_text + table_text + column_text + text

        elif detected_null and not detected_negative:
            detected_null = False
            final_text = initial_text + text

        elif detected_negative and not detected_null:
            detected_negative = False
            final_text = initial_text + table_text + column_text


    if final_text != "":
        text = final_text + " Please manually check database for the errors."
        recipient = "hpalcich@MicroTekFinishing.com"
        subject = "Automated Message: Excel_To_Database.py Database Integrity Error"
        send_mail_via_com(text, subject, recipient)


# Function to send automated emails from the local Outlook account
# The text, subject and recipient can all be modified for the unique error or message that is required.

def send_mail_via_com(text, subject, recipient):
    o = win32.Dispatch("Outlook.Application")

    Msg = o.CreateItem(0)
    Msg.To = recipient

    # Msg.CC = "drichards@MicroTekFinishing.com"

    Msg.Subject = subject
    Msg.Body = text

    Msg.Send()


def check_for_table(partname, parameters, database):

    try:
        conn = sqlite3.connect(database)
        c = conn.cursor()
        tablename = partname
        command = 'CREATE TABLE IF NOT EXISTS Part_' + tablename
        columnnames = parameters
        valuenames = " ("

        case = 1  # Determines whether the excel file contains only before, after or both before and after measurements
        # 1 = only Before, 2 = both, 3 = only After
        if columnnames[1][0].find('Before') != -1 and columnnames[2][0].find('After') != -1:
            case = 2
        elif columnnames[1][0].find('Before') == -1 and columnnames[2][0].find('After') != -1:
            case = 3


        for i in range(0, len(columnnames)):
            if case == 1 or case == 3:
                if i == 0 or 1 or 2 or 3:
                    valuenames = valuenames + columnnames[i][0].replace(" ", "_") + " " + "text, "
                    valuenames = valuenames.replace("/", "_")
                else:
                    valuenames = valuenames + columnnames[i][0].replace(" ", "_") + " " + "real, "
            elif case == 2:
                if i == 0 or 1 or 2 or 3 or 4:
                    valuenames = valuenames + columnnames[i][0].replace(" ", "_") + " " + "text, "
                    valuenames = valuenames.replace("/", "_")
                else:
                    valuenames = valuenames + columnnames[i][0].replace(" ", "_") + " " + "real, "



        valuenames = valuenames[:-2] + ");"
        valuenames = valuenames.replace("-", "_")
        command = command + valuenames
        c.execute(command)
        conn.commit()
        c.close()
    except sqlite3.OperationalError:
        text = "This is an automated email alerting you that the Excel_To_Database.py file had" \
               " an error in the check_for_table function. The database is most likely being used and locked. Save changes" \
               "to database and close the file. Then rerun the data."
        recipient = "hpalcich@MicroTekFinishing.com"
        subject = "Automated Message: Excel_To_Database.py check_for_table function Error"
        send_mail_via_com(text, subject, recipient)


def insert_data(partname, Data, DBlinks, i, database):
    conn = sqlite3.connect(database)
    c = conn.cursor()
    tablename = partname
    command = 'INSERT INTO Part_' + tablename + ' VALUES'
    x = i  # Index variable for looping through Part Data files

    nrows = Data.shape[0]
    ncols = Data.shape[1]  # get number of rows and columns of Data

    values = " ("
    for i in range(0, ncols):  # Construct SQL Query to input data
        values = values + "?,"


    values = values[:-1] + ")"
    command = command + values
    for i in range(0, nrows):  # Input each row into DB
        c.execute(command, Data.iloc[i, :])
        conn.commit()

    c.close()

    wbk = openpyxl.load_workbook(DBlinks.iat[x, 0])
    wks = wbk['Overall Record']
    wks.delete_rows(5, 4 + nrows)

    # write the column back into the excel file and close it
    failure = True
    attempts = 0
    while failure:  # keep trying to save and close workbook
        try:
            wbk.save(DBlinks.iat[x, 0])
            wbk.close()
            failure = False
        except PermissionError:  # if error arises from workbook being opened wait and try again
            if attempts < 2:
                time.sleep(180)  # if an errors occurs, wait 3 mins to try again.
                attempts = attempts + 1
                print('attempt ', attempts)
                continue
            else:  # if errors are persistent then just stop and send an email letting someone know


                text = "This is an automated email alerting you that the Excel_To_Database.py file can" \
                       " not save the " + DBlinks.iat[x, 0] + " file. Most likely someone has it open."
                recipient = "hpalcich@MicroTekFinishing.com"
                subject = "Automated Message: Excel_To_Database.py Workbook Save Error"
                send_mail_via_com(text, subject, recipient)
                failure = False


if __name__ == "__main__":
    # execute only if run as a script
    main()